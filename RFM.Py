import pandas as pd
import numpy as np
from datetime import datetime
import streamlit as st
from sklearn.preprocessing import StandardScaler
import plotly.graph_objs as go
from plotly.subplots import make_subplots
import plotly.express as px
import io
from openpyxl import Workbook
from openpyxl.styles import Font

def calculate_rfm(df, customer_col, date_col, order_col, amount_col):
    # Ensure 'date' is in datetime format
    df[date_col] = pd.to_datetime(df[date_col])
    
    # Calculate Recency
    max_date = df[date_col].max()
    df['recency'] = (max_date - df[date_col]).dt.days
    
    # Calculate Frequency
    df['frequency'] = df.groupby(customer_col)[order_col].transform('count')
    
    # Calculate Monetary
    df['monetary'] = df.groupby(customer_col)[amount_col].transform('sum')
    
    # Get the most recent record for each customer
    rfm = df.groupby(customer_col).agg({
        'recency': 'min',
        'frequency': 'max',
        'monetary': 'sum'
    }).reset_index()
    
    # Assign scores
    for col in ['recency', 'frequency', 'monetary']:
        col_label = col[0].upper()
        if col == 'recency':
            rfm[col_label] = pd.qcut(rfm[col], q=5, labels=range(5, 0, -1))
        else:
            rfm[col_label] = pd.qcut(rfm[col], q=5, labels=range(1,6))
    
    rfm['RFM_Score'] = rfm['R'].astype(str) + rfm['F'].astype(str) + rfm['M'].astype(str)
    
    # Calculate overall score
    rfm['RFM_Overall'] = rfm['R'].astype(int) + rfm['F'].astype(int) + rfm['M'].astype(int)
    
    # Assign clusters based on overall score
    #rfm['Cluster'] = pd.qcut(rfm['RFM_Overall'], q=5, labels=range(5))
    
    return rfm

#def plot_cluster_distributions(rfm):
    metrics = ['recency', 'frequency', 'monetary']
    clusters = sorted(rfm['Cluster'].unique())
    
    fig = make_subplots(rows=len(clusters), cols=3, 
                        subplot_titles=[f"{metric.capitalize()} Distribution" for metric in metrics for _ in clusters],
                        vertical_spacing=0.05)
    
    for i, cluster in enumerate(clusters, 1):
        cluster_data = rfm[rfm['Cluster'] == cluster]
        for j, metric in enumerate(metrics, 1):
            fig.add_trace(
                go.Box(y=cluster_data[metric], name=f'Cluster {cluster}'),
                row=i, col=j
            )
    
    fig.update_layout(height=300*len(clusters), width=1000, 
                      title_text="Cluster Distributions (0: Worst, 4: Best)")
    
    return fig

def get_segment_label(row):
    if row['R'] == 4 and row['F'] == 4 and row['M'] == [4,5]:
        return 'Champions'
    elif row['R'] in [3] and row['F'] in [4] and row['M'] in [4, 5]:
        return 'Loyal Customers'
    elif row['R'] in [3, 4] and row['F'] in [2, 3] and row['M'] in [3, 4]:
        return 'Potential Loyalists'
    elif row['R'] == [4] and row['F'] in [1] and row['M'] in [1, 2]:
        return 'New Customers'
    elif row['R'] in [1] and row['F'] in [3,4] and row['M'] in [4, 5]:
        return 'At-Risk'
    elif row['R'] in [1,2,3] and row['F'] in [1, 2] and row['M'] in [1, 2]:
        return 'Hibernating'
    elif row['R'] in [2] and row['F'] in [3, 4] and row['M'] in [3, 4]:
        return 'Can’t Lose Them'
    else:
        return 'Other'

def get_tier_label(row):
    total_rfm = row['R'] + row['F'] + row['M']
    
    if total_rfm in [10, 11, 12]:
        return 'Platinum'
    elif total_rfm in [9, 8, 7]:
        return 'Gold'
    elif total_rfm in [6, 5, 4]:
        return 'Silver'
    elif total_rfm in [3, 2, 1]:
        return 'Bronze'
    elif total_rfm in [13, 14, 15]:
        return 'VIP'
    else:
        return 'Other'

def get_recommendations(segment):
    recommendations = {
        'Best Customers': "Reward these customers. They can become advocates.",
        'Loyal Customers': "Offer loyalty rewards and personalized communication.",
        'Potential Churners': "Re-engagement campaign. Offer special promotions.",
        'At Risk': "Targeted win-back campaign. Understand their needs.",
        'Lost Customers': "Reactivation campaign. Special offers to win them back."
    }
    return recommendations.get(segment, "Customize approach based on individual RFM scores.")

def convert_df_to_csv(df):
    return df.to_csv(index=False).encode('utf-8')

def create_excel_for_tableau(df, rfm_result):
    # Merge the original dataframe with RFM results
    full_data = df.merge(rfm_result[['Customer_ID', 'Tier', 'R', 'F', 'M']], on='Customer_ID', how='left')
    
    # Create a BytesIO object to hold the Excel file
    excel_buffer = io.BytesIO()
    
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RFM Analysis Data"
    
    # Write headers
    headers = list(full_data.columns)
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Write data
    for row, data in enumerate(full_data.itertuples(index=False), start=2):
        for col, value in enumerate(data, start=1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0.00'  # Format numbers with two decimal places
            else:
                cell.value = str(value)
    
    # Save the workbook to the BytesIO object
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def main():
    st.title("MVP Enhanced RFM Analysis App")
    
    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")
    
    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)
        
        st.subheader("Column Mapping")
        st.write("Please select the columns from your Excel file that correspond to the required RFM analysis fields:")
        
        customer_col = st.selectbox("Select the customer ID column", df.columns)
        date_col = st.selectbox("Select the date column", df.columns)
        order_col = st.selectbox("Select the order ID column", df.columns)
        amount_col = st.selectbox("Select the total amount column", df.columns)
        
        if st.button("Perform RFM Analysis"):
            rfm_result = calculate_rfm(df, customer_col, date_col, order_col, amount_col)
            
            rfm_result['Segment'] = rfm_result.apply(get_segment_label, axis=1)

            rfm_result['Tier'] = rfm_result.apply(get_tier_label, axis=1)
            
            st.subheader("RFM Analysis Results")
            st.write(rfm_result)
            
            #rfm_extract = rfm_result[['R', 'F', 'M', 'Cluster', 'Segment']]
            # Mostrar la tabla con las columnas seleccionadas
            #st.write(rfm_extract)
            #rfm_grouped = rfm_extract.groupby(['R', 'F', 'M', 'Cluster', 'Segment']).size().reset_index(name='Número de Clientes')
            #st.write(rfm_grouped)
            
            #st.subheader("Customer Segmentation")
            #fig = go.Figure(data=[go.Scatter3d(
            #    x=rfm_result['R'],
            #    y=rfm_result['F'],
            #    z=rfm_result['M'],
            #    mode='markers',
            #    marker=dict(
            #        size=5,
            #        color=rfm_result['Tier'],
            #        colorscale='Viridis',
            #        opacity=0.8
            #    ),
            #    text=rfm_result[customer_col],
            #    hoverinfo='text'
            #)])
            #fig.update_layout(scene = dict(
            #    xaxis_title='Recency',
            #    yaxis_title='Frequency',
            #    zaxis_title='Monetary'
            #), width=700, margin=dict(r=20, b=10, l=10, t=10))
            #st.plotly_chart(fig)

            # Asegúrate de que rfm_result tiene los datos correctos
            segment_counts = rfm_result['Tier'].value_counts().reindex(['Platinum', 'Gold', 'Silver', 'Bronze', 'VIP', 'Other'], fill_value=0)

            # Define custom colors for each segment
            colors = {
                'Platinum': '#E5E4E2',  # Platinum color
                'Gold': '#FFD700',      # Gold color
                'Silver': '#C0C0C0',    # Silver color
                'Bronze': '#CD7F32',    # Bronze color
                'VIP': '#4B0082',       # Indigo for VIP
                'Other': '#808080'      # Gray for Other
            }

            st.subheader("Segment Distribution")
            fig = go.Figure(data=[go.Pie(labels=segment_counts.index, values=segment_counts.values,marker=dict(colors=[colors[segment] for segment in segment_counts.index]))])
            fig.update_layout(title='Customer Segments')
            st.plotly_chart(fig)

            #st.subheader("Cluster Distributions")
            #fig = plot_cluster_distributions(rfm_result)
            #st.plotly_chart(fig)

            # Create a chart for each customer tier
            tiers = ['VIP', 'Platinum', 'Gold', 'Silver', 'Bronze']

            result_df = df.merge(rfm_result[['Customer_ID', 'Tier']], on='Customer_ID', how='left')

            #tier_data = rfm_result[['Tier','Customer_ID','monetary','InvoiceDate']]

            for tier in tiers:
                # Filter data for the current tier
                tier_df = result_df[result_df['Tier'] == tier]
                
                if not tier_df.empty:
                    # Group by week and calculate weekly sales and distinct customer count
                    weekly_data = tier_df.groupby(pd.Grouper(key='InvoiceDate', freq='W-MON')).agg({
                        'Total_amount': 'sum',
                        'Customer_ID': 'nunique'
                    }).reset_index()

                    # Create the figure
                    fig = go.Figure()

                    # Add the bar trace for distinct customer count at the bottom
                    fig.add_trace(
                        go.Bar(
                            x=weekly_data['InvoiceDate'],
                            y=weekly_data['Customer_ID'],
                            name='Distinct Customers',
                            marker_color='lightblue',
                            yaxis='y2'
                        )
                    )

                    # Add the line trace for weekly sales
                    fig.add_trace(
                        go.Scatter(
                            x=weekly_data['InvoiceDate'],
                            y=weekly_data['Total_amount'],
                            name='Weekly Sales',
                            mode='lines',
                            line=dict(color='blue'),
                            yaxis='y'
                        )
                    )

                    # Set titles and labels
                    fig.update_layout(
                        title_text=f"{tier} Customers: Weekly Sales and Customer Count",
                        xaxis_title="Week",
                        yaxis=dict(
                            title="Weekly Sales",
                            side="left",
                            fixedrange=False  # Allow y-axis to be scaled
                        ),
                        yaxis2=dict(
                            title="Distinct Customer Count",
                            side="right",
                            overlaying="y",
                            fixedrange=False,  # Allow y2-axis to be scaled
                            showgrid=False
                        ),
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=1.02,
                            xanchor="right",
                            x=1
                        )
                    )

                    # Enable y-axis modification with mouse
                    fig.update_layout(dragmode='zoom', hovermode='x unified')
                    fig.update_xaxes(rangeslider_visible=True)

                    # Show the figure in Streamlit
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.write(f"No data available for {tier} tier")
            
            st.subheader("Top 10 Customers by RFM Score")
            top_customers = rfm_result.sort_values('Tier', ascending=False).head(10)
            st.write(top_customers)
                        
            # Calculate total amount for each product
            product_totals = result_df.groupby('StockCode').agg({
                'Total_amount': 'sum',
                'Description': 'first'  # Assuming 'Description' is the product name
            }).sort_values('Total_amount', ascending=False)

            # Get the top 5 products
            top_5_products = product_totals.head(5).reset_index()

            st.write("Top 5 Products by Total Amount:")
            st.write(top_5_products)

            # Create subplots for bar charts
            fig = make_subplots(rows=3, cols=2, 
                                subplot_titles=[f"{desc} (Code: {code})" for code, desc in zip(top_5_products['StockCode'], top_5_products['Description'])],
                                vertical_spacing=0.2,
                                horizontal_spacing=0.1)

            # Get unique tiers from the data
            tiers = result_df['Tier'].unique()

            # Generate a color map for all tiers
            color_scale = px.colors.qualitative.Plotly
            color_map = {tier: color_scale[i % len(color_scale)] for i, tier in enumerate(tiers)}


            # Create bar charts for each product
            for i, row in enumerate(top_5_products.itertuples(), 1):
                stock_code = row.StockCode
                product_df = result_df[result_df['StockCode'] == stock_code]
                
                tier_data = product_df.groupby('Tier').agg({
                    'Total_amount': ['sum']
                }).reset_index()
                tier_data.columns = ['Tier', 'Total']
                
                row_idx = (i - 1) // 2 + 1
                col_idx = (i - 1) % 2 + 1
                
                if not tier_data.empty:
                    
                    # Bar chart for total amount
                    fig.add_trace(
                        go.Bar(x=tier_data['Tier'], y=tier_data['Total'], name='Total', 
                            marker_color=[color_map.get(tier, 'gray') for tier in tier_data['Tier']], 
                            text=tier_data['Total'].round(2), textposition='auto', opacity=0.7),
                        row=row_idx, col=col_idx
                    )
                    
                    fig.update_xaxes(title_text="Customer Tier", row=row_idx, col=col_idx)
                    fig.update_yaxes(title_text="Amount", row=row_idx, col=col_idx)
                else:
                    st.write(f"No data available for product: {row.Description}")

            # Update layout
            fig.update_layout(
                height=2200,  # Increased height
                width=1200, 
                title_text="Median and Total Amount by Customer Tier for Top 5 Products",
                showlegend=True, 
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=0.7,  # Moved legend up
                    xanchor="center",
                    x=0.5,  # Centered legend
                    bgcolor="rgba(255,255,255,0.8)",  # Semi-transparent background
                ),
                margin=dict(t=150)  # Increased top margin for legend
            )

            # Adjust subplot titles to be more prominent
            for i in fig['layout']['annotations']:
                i['font'] = dict(size=12, color='darkblue')
                i['y'] = i['y'] + 0.03  # Move subplot titles up slightly

            # Show the figure in Streamlit
            st.plotly_chart(fig)

            excel_file = create_excel_for_tableau(df, rfm_result)
            st.write(df)
            st.write(rfm_result)
            st.write(excel_file)
            
            # Add a download button for Excel file
            st.download_button(label='Download Current Result',
                                data=excel_file ,
                                file_name= 'Segementation_customer.xlsx')
    else:
        st.info("Please upload an Excel file to perform RFM analysis.")

if __name__ == "__main__":
    main()

#CODE